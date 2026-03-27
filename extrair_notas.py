"""
Script de Extração de Notas de Corretagem - Genial CCTVM
Extrai dados financeiros dos PDFs e gera planilha Excel anual consolidada.

Estrutura da planilha:
  - Uma aba por mês (ex: "Março", "Abril" ...)
  - Última aba: "TOTAL ANUAL" com totalizador de todos os meses
"""

import pdfplumber
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import re
import os
from collections import defaultdict

# ── Mapeamento de mês (número → nome em português) ──────────────────────────
MESES_PT = {
    1: 'Janeiro', 2: 'Fevereiro', 3: 'Março',    4: 'Abril',
    5: 'Maio',    6: 'Junho',     7: 'Julho',     8: 'Agosto',
    9: 'Setembro',10: 'Outubro',  11: 'Novembro', 12: 'Dezembro',
}

# ── Colunas exibidas na planilha ─────────────────────────────────────────────
COLUMNS = [
    'Nr. Nota',
    'Data Pregão',
    'Valor dos Negócios',
    'IRRF',
    'IRRF Day Trade (Projeção)',
    'Taxa Operacional',
    'Taxa Registro BM&F',
    'Taxas BM&F (emol+f.gar)',
    'Outros Custos',
    'ISS',
    'Ajuste Day Trade',
    'Total das Despesas',
    'Outros',
    'Total Líquido da Nota',
]

# Índice da última coluna de texto (não somável)
N_TEXT_COLS = 2   # Nr. Nota + Data Pregão


# ── Helpers ──────────────────────────────────────────────────────────────────

def parse_float(val_str):
    """Converte string numérica brasileira ('1.234,56') para float."""
    val_str = val_str.strip().replace('.', '').replace(',', '.')
    try:
        return float(val_str)
    except ValueError:
        return 0.0


def extract_values_row(line):
    """
    Extrai lista de floats sinalizados de uma linha de resumo financeiro.
    D → negativo  |  C (ou sem marcação) → positivo

    Formatos:
    1. D/C único no final → aplica-se a todos: '0,00 3,94 0,00 8,64 4,86 D'
    2. D/C inline por valor: '0,00 C 408,00 C 13,50 D'
    """
    tokens = re.findall(r'([\d\.]+,\d+)\s*([DC]?)', line)
    has_inline_dc = any(dc for _, dc in tokens[:-1])

    if has_inline_dc:
        return [parse_float(n) * (-1 if dc == 'D' else 1) for n, dc in tokens]
    else:
        row_dc = tokens[-1][1] if tokens else ''
        sign = -1 if row_dc == 'D' else 1
        return [parse_float(n) * sign for n, _ in tokens]


def extract_nota_data(page_text, nota_num, data_pregao):
    """Extrai os campos financeiros de uma página de fechamento de nota."""
    lines = page_text.split('\n')

    data = {
        'Nr. Nota': nota_num,
        'Data Pregão': data_pregao,
        'Valor dos Negócios': 0.0,
        'IRRF': 0.0,
        'IRRF Day Trade (Projeção)': 0.0,
        'Taxa Operacional': 0.0,
        'Taxa Registro BM&F': 0.0,
        'Taxas BM&F (emol+f.gar)': 0.0,
        'Outros Custos': 0.0,
        'ISS': 0.0,
        'Ajuste de Posição': 0.0,
        'Ajuste Day Trade': 0.0,
        'Total das Despesas': 0.0,
        'Outros': 0.0,
        'IRRF Corretagem': 0.0,
        'Total Conta Investimento': 0.0,
        'Total Conta Normal': 0.0,
        'Total Líquido (#)': 0.0,
        'Total Líquido da Nota': 0.0,
    }

    for i, line in enumerate(lines):
        if 'Venda disponível' in line and 'Valor dos negócios' in line:
            if i + 1 < len(lines):
                vals = extract_values_row(lines[i + 1])
                if len(vals) >= 5:
                    data['Valor dos Negócios'] = vals[4]

        elif 'IRRF' in line and 'Taxa registro BM' in line and 'Taxas BM' in line:
            if i + 1 < len(lines):
                vals = extract_values_row(lines[i + 1])
                if len(vals) >= 5:
                    data['IRRF']                      = vals[0]
                    data['IRRF Day Trade (Projeção)'] = vals[1]
                    data['Taxa Operacional']           = vals[2]
                    data['Taxa Registro BM&F']         = vals[3]
                    data['Taxas BM&F (emol+f.gar)']   = vals[4]

        elif '+ Outros Custos' in line and 'Total das despesas' in line:
            if i + 1 < len(lines):
                vals = extract_values_row(lines[i + 1])
                if len(vals) >= 5:
                    data['Outros Custos']      = vals[0]
                    data['ISS']                = vals[1]
                    data['Ajuste de Posição']  = vals[2]
                    data['Ajuste Day Trade']   = vals[3]
                    data['Total das Despesas'] = vals[4]

        elif 'Outros' in line and 'Total líquido da nota' in line and 'IRRF Corretagem' in line:
            if i + 1 < len(lines):
                vals = extract_values_row(lines[i + 1])
                if len(vals) >= 6:
                    data['Outros']                   = vals[0]
                    data['IRRF Corretagem']          = vals[1]
                    data['Total Conta Investimento'] = vals[2]
                    data['Total Conta Normal']       = vals[3]
                    data['Total Líquido (#)']        = vals[4]
                    data['Total Líquido da Nota']    = vals[5]

    return data


def extract_header(text):
    """Extrai Nr. nota e Data pregão do cabeçalho da nota."""
    nota_num = data_pregao = None
    lines = text.split('\n')
    for i, line in enumerate(lines):
        if 'Nr. nota' in line and 'Data pregão' in line:
            if i + 1 < len(lines):
                parts = lines[i + 1].strip().split()
                if len(parts) >= 3:
                    nota_num   = parts[0]
                    data_pregao = parts[2]
            break
    return nota_num, data_pregao


def process_pdf(pdf_path):
    """Processa um PDF e retorna lista de dicts, um por nota."""
    notas = []
    with pdfplumber.open(pdf_path) as pdf:
        current_nota_num = None
        current_data_pregao = None

        for page in pdf.pages:
            text = page.extract_text() or ""
            nota_num, data_pregao = extract_header(text)

            if nota_num and nota_num != current_nota_num:
                current_nota_num    = nota_num
                current_data_pregao = data_pregao

            is_closing = (
                'Venda disponível'    in text
                and 'Total líquido da nota' in text
                and 'CONTINUA...'         not in text
            )

            if is_closing and current_nota_num:
                nota_data = extract_nota_data(text, current_nota_num, current_data_pregao)
                notas.append(nota_data)

    return notas


# ── Estilos compartilhados ───────────────────────────────────────────────────

def make_styles():
    thin = Side(border_style="thin", color="BFBFBF")
    return {
        'header_fill':  PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid"),
        'header_font':  Font(name='Calibri', bold=True, color="FFFFFF", size=10),
        'alt_fill':     PatternFill(start_color="D6E4F0", end_color="D6E4F0", fill_type="solid"),
        'total_fill':   PatternFill(start_color="F2C94C", end_color="F2C94C", fill_type="solid"),
        'bold_font':    Font(name='Calibri', bold=True, size=10),
        'normal_font':  Font(name='Calibri', size=10),
        'center':       Alignment(horizontal='center', vertical='center', wrap_text=True),
        'right':        Alignment(horizontal='right',  vertical='center'),
        'border':       Border(left=thin, right=thin, top=thin, bottom=thin),
        'num_fmt':      '#,##0.00',
    }


def write_sheet(ws, notas, st):
    """Escreve os dados de um mês em uma aba worksheet."""
    # Cabeçalho
    for col_idx, col_name in enumerate(COLUMNS, 1):
        cell = ws.cell(row=1, column=col_idx, value=col_name)
        cell.font      = st['header_font']
        cell.fill      = st['header_fill']
        cell.alignment = st['center']
        cell.border    = st['border']
    ws.row_dimensions[1].height = 40
    ws.freeze_panes = "A2"

    # Dados
    for row_idx, nota in enumerate(notas, 2):
        fill = st['alt_fill'] if row_idx % 2 == 0 else PatternFill()
        for col_idx, col_name in enumerate(COLUMNS, 1):
            val  = nota.get(col_name, '')
            cell = ws.cell(row=row_idx, column=col_idx, value=val)
            cell.border = st['border']
            cell.fill   = fill
            cell.font   = st['normal_font']
            if col_idx <= N_TEXT_COLS:
                cell.alignment = st['center']
            else:
                cell.alignment     = st['right']
                cell.number_format = st['num_fmt']

    # Linha de totais do mês
    total_row = len(notas) + 2
    for col_idx in range(1, len(COLUMNS) + 1):
        cell        = ws.cell(row=total_row, column=col_idx)
        cell.fill   = st['total_fill']
        cell.border = st['border']
        cell.font   = st['bold_font']
        if col_idx == 1:
            cell.value     = 'TOTAL'
            cell.alignment = st['center']
        elif col_idx == 2:
            cell.value     = '---'
            cell.alignment = st['center']
        else:
            col_letter         = get_column_letter(col_idx)
            cell.value         = f"=SUM({col_letter}2:{col_letter}{total_row - 1})"
            cell.number_format = st['num_fmt']
            cell.alignment     = st['right']

    # Larguras
    for col_idx in range(1, len(COLUMNS) + 1):
        ws.column_dimensions[get_column_letter(col_idx)].width = 12 if col_idx <= N_TEXT_COLS else 20

    return total_row   # linha de total (útil para referência externa)


def write_annual_sheet(ws, month_sheets, st):
    """
    Aba final com totalizador anual.
    Cada linha representa um mês; última linha soma tudo.
    month_sheets: list of (sheet_name, ws_ref, total_row)
    """
    # Cabeçalho — sem Nr.Nota / Data Pregão nas colunas de soma
    annual_cols = ['Mês'] + COLUMNS[N_TEXT_COLS:]

    for col_idx, col_name in enumerate(annual_cols, 1):
        cell = ws.cell(row=1, column=col_idx, value=col_name)
        cell.font      = st['header_font']
        cell.fill      = st['header_fill']
        cell.alignment = st['center']
        cell.border    = st['border']
    ws.row_dimensions[1].height = 40
    ws.freeze_panes = "A2"

    # Uma linha por mês referenciando a linha TOTAL de cada aba
    for row_idx, (sheet_name, ref_ws, total_row) in enumerate(month_sheets, 2):
        fill = st['alt_fill'] if row_idx % 2 == 0 else PatternFill()

        # Coluna 1: nome do mês
        cell            = ws.cell(row=row_idx, column=1, value=sheet_name)
        cell.fill       = fill
        cell.border     = st['border']
        cell.font       = st['bold_font']
        cell.alignment  = st['center']

        # Colunas numéricas: referencia a célula TOTAL da aba do mês
        for col_idx, col_name in enumerate(COLUMNS[N_TEXT_COLS:], 2):
            src_col    = COLUMNS.index(col_name) + 1          # col na aba do mês
            src_letter = get_column_letter(src_col)
            cell = ws.cell(
                row=row_idx, column=col_idx,
                value=f"='{sheet_name}'!{src_letter}{total_row}"
            )
            cell.fill          = fill
            cell.border        = st['border']
            cell.font          = st['normal_font']
            cell.alignment     = st['right']
            cell.number_format = st['num_fmt']

    # Linha TOTAL ANUAL
    total_row_annual = len(month_sheets) + 2
    annual_total_fill = PatternFill(start_color="2E7D32", end_color="2E7D32", fill_type="solid")
    white_bold = Font(name='Calibri', bold=True, color="FFFFFF", size=11)

    for col_idx in range(1, len(annual_cols) + 1):
        cell        = ws.cell(row=total_row_annual, column=col_idx)
        cell.fill   = annual_total_fill
        cell.border = st['border']
        cell.font   = white_bold
        if col_idx == 1:
            cell.value     = 'TOTAL ANUAL'
            cell.alignment = st['center']
        else:
            col_letter         = get_column_letter(col_idx)
            cell.value         = f"=SUM({col_letter}2:{col_letter}{total_row_annual - 1})"
            cell.number_format = st['num_fmt']
            cell.alignment     = st['right']

    # Larguras
    ws.column_dimensions['A'].width = 14
    for col_idx in range(2, len(annual_cols) + 1):
        ws.column_dimensions[get_column_letter(col_idx)].width = 20


def save_to_excel(notas_por_mes, ano, output_path):
    """
    notas_por_mes: dict { (ano, mes): [lista de notas] }
    Cria uma aba por mês + aba de totais anuais.
    """
    wb = openpyxl.Workbook()
    wb.remove(wb.active)   # remove planilha padrão vazia
    st = make_styles()

    month_sheets = []   # (sheet_name, ws, total_row)

    # Ordena meses cronologicamente
    for (y, m) in sorted(notas_por_mes.keys()):
        if y != ano:
            continue
        sheet_name = MESES_PT.get(m, f"Mês {m:02d}")
        ws = wb.create_sheet(title=sheet_name)
        total_row = write_sheet(ws, notas_por_mes[(y, m)], st)
        month_sheets.append((sheet_name, ws, total_row))

    # Aba de totais anuais
    ws_annual = wb.create_sheet(title=f"TOTAL {ano}")
    write_annual_sheet(ws_annual, month_sheets, st)

    wb.save(output_path)
    print(f"Planilha salva em: {output_path}")


# ── Main ─────────────────────────────────────────────────────────────────────

def main():
    resources_dir = r"c:\Repo\Planilha Acoes\resources"

    # Processa todos os PDFs disponíveis
    files_to_process = sorted(f for f in os.listdir(resources_dir) if f.endswith('.pdf'))
    print(f"Arquivos encontrados: {files_to_process}\n")

    notas_por_mes = defaultdict(list)   # {(ano, mes): [notas]}
    anos = set()

    for filename in files_to_process:
        pdf_path = os.path.join(resources_dir, filename)
        print(f"Processando: {filename}")
        notas = process_pdf(pdf_path)
        print(f"  → {len(notas)} notas extraídas")

        for nota in notas:
            # Extrai mês/ano da data pregão (formato dd/mm/yyyy)
            try:
                partes = nota['Data Pregão'].split('/')
                mes = int(partes[1])
                ano = int(partes[2])
            except (IndexError, ValueError, AttributeError):
                print(f"  ⚠ Data inválida em nota {nota.get('Nr. Nota')}: {nota.get('Data Pregão')}")
                continue

            notas_por_mes[(ano, mes)].append(nota)
            anos.add(ano)

    print(f"\nResume: {sum(len(v) for v in notas_por_mes.values())} notas no total")
    for (ano, mes), lst in sorted(notas_por_mes.items()):
        print(f"  {MESES_PT.get(mes, mes)}/{ano}: {len(lst)} notas")

    # Gera uma planilha por ano encontrado
    for ano in sorted(anos):
        output_path = os.path.join(r"c:\Repo\Planilha Acoes", f"notas_corretagem_{ano}.xlsx")
        save_to_excel(notas_por_mes, ano, output_path)


if __name__ == '__main__':
    main()
